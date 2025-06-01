# niveaux/views.py
from django.http import HttpResponse
from rest_framework import generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework.filters import SearchFilter, OrderingFilter
from openpyxl import load_workbook
from io import BytesIO
from .models import Niveau
from .serializers import NiveauSerializer
from professeurs.views import PaginatorPage

class NiveauListCreateView(generics.ListCreateAPIView):
    queryset = Niveau.objects.all()
    serializer_class = NiveauSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PaginatorPage
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['nom']
    ordering_fields = ['nom']

class NiveauDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Niveau.objects.all()
    serializer_class = NiveauSerializer
    permission_classes = [IsAuthenticated]

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def niveau_import_excel(request):
    # Vérifier si un fichier a été envoyé
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    
    # Vérifier l'extension du fichier
    if not file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'Invalid file format. Please upload an Excel file'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Charger le fichier Excel
        workbook = load_workbook(filename=BytesIO(file.read()))
        sheet = workbook.active
        
        # Vérifier l'en-tête (attendu : "nom")
        header = [cell.value for cell in sheet[1]]
        if header != ['nom']:
            return Response({'error': 'Invalid header. Expected: nom'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Traiter les lignes
        results = []
        errors = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nom = row[0]
            
            if not nom:
                errors.append({'row': row[0], 'error': 'Empty name'})
                continue
            
            # Vérifier l'unicité
            if Niveau.objects.filter(nom=nom).exists():
                errors.append({'row': nom, 'error': 'Name already exists'})
                continue
            
            # Créer le niveau
            try:
                niveau = Niveau.objects.create(nom=nom)
                results.append(NiveauSerializer(niveau).data)
            except Exception as e:
                errors.append({'row': nom, 'error': str(e)})
        
        response = {
            'imported': results,
            'errors': errors,
            'message': f'{len(results)} levels imported, {len(errors)} errors'
        }
        return Response(response, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response({'error': f'Error processing file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)